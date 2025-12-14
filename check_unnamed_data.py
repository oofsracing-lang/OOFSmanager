import pandas as pd
import openpyxl

# Open the Excel file
wb = openpyxl.load_workbook('XML Ingest - DEV.xlsx')
ws = wb['Drivers']

# Get headers
headers = [cell.value for cell in ws[1]]

# Get first 5 data rows to see what's in the unnamed columns
print("Sample data from unnamed columns (28-38):")
for row_idx in range(2, 7):  # rows 2-6
    row = ws[row_idx]
    driver_name = row[0].value
    print(f"\n{driver_name}:")
    for col_idx in range(28, 38):  # columns 28-37
        value = row[col_idx].value
        print(f"  Column {col_idx}: {value}")
